import os
import platform
import shutil
import smtplib
import subprocess
import threading
import tkinter as tk
import re
import webbrowser
from collections import Counter
from datetime import date, datetime, timedelta
from email.mime.text import MIMEText
from tkinter import ttk, messagebox, filedialog, simpledialog
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from openpyxl import Workbook, load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from tkcalendar import DateEntry, Calendar
import time

EXCEL_FILE = os.path.join(os.path.dirname(__file__), "real_estate_leads.xlsx")
PDF_FILE = os.path.join(os.path.dirname(__file__), "leads.pdf")
BACKUP_FOLDER = os.path.join(os.path.dirname(__file__), "backups")
TASKS_FILE = os.path.join(os.path.dirname(__file__), "tasks.txt")
INVOICE_FILE = os.path.join(os.path.dirname(__file__), "invoices.xlsx")
INVOICE_PDF = os.path.join(os.path.dirname(__file__), "invoice.pdf")
INVOICE_HTML = os.path.join(os.path.dirname(__file__), "invoice.html")
LOGO_PATH = os.path.join(os.path.dirname(__file__), "logo.png")

# ---------------------------- Utility Functions ----------------------------
def is_duplicate_lead(name, phone, email):
    leads = get_leads()
    for lead in leads:
        if lead.get("Phone") == phone or lead.get("Email") == email:
            return True
    return False


def get_leads():

    if not os.path.exists(EXCEL_FILE):
        return []
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    return [dict(zip([cell.value for cell in ws[1]], [cell.value for cell in row])) for row in ws.iter_rows(min_row=2)]

def write_leads(leads):
    global lead
    wb = Workbook()
    ws = wb.active
    headers = ["Name", "Phone", "Email", "Source", "Property", "Status", "Follow-up", "Notes"]
    ws.append(headers)
    for lead in leads:
        ws.append([lead.get(h, '') for h in headers])
    wb.save(EXCEL_FILE)
    backup_excel()
    if is_duplicate_lead(lead["Name"], lead["Phone"], lead["Email"]):
        messagebox.showwarning("Duplicate", "A lead with this phone or email already exists.")
        return


def backup_excel():
    try:
        if not os.path.exists(BACKUP_FOLDER):
            os.makedirs(BACKUP_FOLDER)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_file = os.path.join(BACKUP_FOLDER, f"backup_{timestamp}.xlsx")
        shutil.copy2(EXCEL_FILE, backup_file)
    except Exception as e:
        print(f"Backup failed: {e}")

def open_backup_folder():
    if not os.path.exists(BACKUP_FOLDER):
        os.makedirs(BACKUP_FOLDER)
    if platform.system() == "Windows":
        os.startfile(BACKUP_FOLDER)
    elif platform.system() == "Darwin":
        subprocess.call(["open", BACKUP_FOLDER])
    else:
        subprocess.call(["xdg-open", BACKUP_FOLDER])

def restore_backup():
    file_path = filedialog.askopenfilename(initialdir=BACKUP_FOLDER, title="Select Backup File", filetypes=[("Excel Files", "*.xlsx")])
    if file_path:
        try:
            shutil.copy2(file_path, EXCEL_FILE)
            messagebox.showinfo("Restored", "Backup restored successfully.")
            search_leads()
        except Exception as e:
            messagebox.showerror("Restore Failed", f"Could not restore backup:\n{e}")

def export_pdf(selected_leads=None):
    leads = selected_leads if selected_leads else get_leads()
    if not leads:
        messagebox.showwarning("No Data", "No leads to export.")
        return
    c = canvas.Canvas(PDF_FILE, pagesize=A4)
    width, height = A4
    y = height - 40
    c.setFont("Helvetica-Bold", 10)
    for lead in leads:
        text = f"{lead['Name']} | {lead['Phone']} | {lead['Email']} | {lead['Status']} | {lead['Follow-up']}"
        c.drawString(30, y, text)
        y -= 20
        if y < 50:
            c.showPage()
            y = height - 40
    c.save()
    os.startfile(PDF_FILE)

def print_pdf():
    selected_items = result_tree.selection()
    if selected_items:
        all_leads = get_leads()
        selected_leads = [all_leads[int(iid)] for iid in selected_items]
        export_pdf(selected_leads)
    elif not os.path.exists(PDF_FILE):
        messagebox.showwarning("Print PDF", "PDF not found. Export it first.")
        return
    try:
        if platform.system() == "Windows":
            os.startfile(PDF_FILE, "print")
        elif platform.system() == "Darwin":
            os.system(f"lp '{PDF_FILE}'")
        else:
            os.system(f"lp '{PDF_FILE}'")
    except Exception as e:
        messagebox.showerror("Print Error", f"Could not print PDF:\n{e}")

def send_whatsapp():
    phone = fields["Phone"].get().strip().replace("-", "")
    name = fields["Name"].get().strip()
    if not phone:
        return
    if phone.startswith("0"):
        phone = "92" + phone[1:]
    url = f"https://wa.me/{phone}?text=Hello%20{name}%2C%20from%20Look-Out%20Real%20Property."
    webbrowser.open(url)

def send_email():
    to = simpledialog.askstring("Send Email", "Recipient Email:")
    msg = MIMEText(f"Hello {fields['Name'].get()},\n\nThis is a message from Look-Out Real Property.")
    msg["Subject"] = "Follow-up"
    msg["From"] = "your_email@gmail.com"
    msg["To"] = to
    try:
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login("your_email@gmail.com", "your_app_password")
        server.sendmail("your_email@gmail.com", to, msg.as_string())
        server.quit()
        messagebox.showinfo("Email", "Email sent successfully.")
    except Exception as e:
        messagebox.showerror("Email Failed", str(e))

def search_leads():
    name_q = search_name_var.get().lower()
    status_q = search_status_var.get()
    leads = get_leads()
    result_tree.delete(*result_tree.get_children())
    for i, l in enumerate(leads):
        if (not name_q or name_q in l["Name"].lower()) and (status_q == "All" or l["Status"] == status_q):
            result_tree.insert("", tk.END, iid=str(i), values=(l["Name"], l["Phone"], l["Email"], l["Source"], l["Property"], l["Status"], l["Follow-up"], l["Notes"]))

def add_lead():
    lead = {key: fields[key].get() for key in fields}
    lead["Notes"] = notes_text.get("1.0", tk.END).strip()
    lead["Follow-up"] = followup_date.get()
    if not lead["Name"] or not lead["Phone"]:
        messagebox.showerror("Error", "Name and Phone are required.")
        return
    leads = get_leads()
    leads.append(lead)
    write_leads(leads)
    messagebox.showinfo("Success", "Lead added.")
    clear_form()
    search_leads()

def load_selected_lead():
    selected = result_tree.selection()
    if not selected:
        return
    index = int(selected[0])
    lead = get_leads()[index]
    for key in fields:
        fields[key].delete(0, tk.END)
        fields[key].insert(0, lead[key])
    followup_date.set_date(lead["Follow-up"])
    notes_text.delete("1.0", tk.END)
    notes_text.insert("1.0", lead["Notes"])
    global selected_lead_index
    selected_lead_index = index

def update_lead():
    if selected_lead_index is None:
        return
    leads = get_leads()
    updated = {key: fields[key].get() for key in fields}
    updated["Notes"] = notes_text.get("1.0", tk.END).strip()
    updated["Follow-up"] = followup_date.get()
    leads[selected_lead_index] = updated
    write_leads(leads)
    messagebox.showinfo("Updated", "Lead updated successfully.")
    clear_form()
    search_leads()

def delete_lead():
    selected = result_tree.selection()
    if not selected:
        return
    index = int(selected[0])
    leads = get_leads()
    if index < len(leads):
        del leads[index]
        write_leads(leads)
        search_leads()
        messagebox.showinfo("Deleted", "Lead deleted.")

def clear_form():
    for entry in fields.values():
        entry.delete(0, tk.END)
    notes_text.delete("1.0", tk.END)
    followup_date.set_date(date.today())


def show_followups_today():
    today_str = date.today().strftime("%Y-%m-%d")
    leads = get_leads()
    today_leads = [l for l in leads if str(l.get("Follow-up")) == today_str]
    if today_leads:
        names = "\n".join([l["Name"] for l in today_leads])
        messagebox.showinfo("Today's Follow-Ups", f"Leads to follow up today:\n\n{names}")

def show_inactive_leads(days_threshold=30):
    leads = get_leads()
    inactive = []
    today = date.today()
    for lead in leads:
        try:
            fup = datetime.strptime(str(lead.get("Follow-up")), "%Y-%m-%d").date()
            if (today - fup).days > days_threshold:
                inactive.append(lead)
        except:
            continue
    if inactive:
        names = "\n".join([f"{l['Name']} (Last: {l['Follow-up']})" for l in inactive])
        messagebox.showinfo("Inactive Leads", f"Leads not contacted in over {days_threshold} days:\n\n{names}")
    else:
        messagebox.showinfo("Inactive Leads", f"No inactive leads older than {days_threshold} days.")



# Tag system
TAGS_FILE = os.path.join(os.path.dirname(__file__), "tags.txt")

def load_tags():
    if os.path.exists(TAGS_FILE):
        with open(TAGS_FILE, "r") as f:
            return [line.strip() for line in f if line.strip()]
    return []

def save_tags(tags):
    with open(TAGS_FILE, "w") as f:
        for tag in tags:
            f.write(tag + "\n")

def assign_tag_to_selected(tag):
    selected = result_tree.selection()
    if not selected:
        return
    index = int(selected[0])
    leads = get_leads()
    leads[index]["Notes"] += f"\n[Tag: {tag}]"
    write_leads(leads)
    search_leads()

def create_tag_menu(frame):
    tags = load_tags()
    tag_menu = ttk.Combobox(frame, values=tags, width=15)
    tag_menu.pack(side="left", padx=5)
    def assign():
        tag = tag_menu.get().strip()
        if tag:
            assign_tag_to_selected(tag)
    ttk.Button(frame, text="Assign Tag", command=assign).pack(side="left", padx=5)
    return tag_menu

# Print Individual Lead
INDIVIDUAL_PDF = os.path.join(os.path.dirname(__file__), "lead_detail.pdf")

def print_selected_lead_detail():
    selected = result_tree.selection()
    if not selected:
        messagebox.showwarning("No Selection", "Please select a lead to print.")
        return
    index = int(selected[0])
    lead = get_leads()[index]
    c = canvas.Canvas(INDIVIDUAL_PDF, pagesize=A4)
    c.setFont("Helvetica-Bold", 12)
    y = 800
    c.drawString(30, y, "Lead Detail Report")
    c.setFont("Helvetica", 10)
    y -= 40
    for key in ["Name", "Phone", "Email", "Source", "Property", "Status", "Follow-up", "Notes"]:
        lines = str(lead.get(key, "")).splitlines()
        for line in lines:
            c.drawString(30, y, f"{key}: {line}")
            y -= 20
        y -= 10
    c.save()
    try:
        if platform.system() == "Windows":
            os.startfile(INDIVIDUAL_PDF, "print")
        else:
            os.system(f"lp '{INDIVIDUAL_PDF}'")
    except Exception as e:
        messagebox.showerror("Print Error", str(e))

# ----------------------------  Invoice Tab (Utility Functions) ----------------------------
def show_invoice_tab(notebook):
    invoice_tab = ttk.Frame(notebook)
    notebook.add(invoice_tab, text="Invoices")

    form_frame = ttk.Frame(invoice_tab)
    form_frame.pack(side="top", fill="x", padx=10, pady=10)

    ttk.Label(form_frame, text="Client:").grid(row=0, column=0, sticky="w")
    client_entry = ttk.Entry(form_frame, width=40)
    client_entry.grid(row=0, column=1, sticky="w", padx=5, pady=5)
    ttk.Label(form_frame, text="Items (format: Description,Qty,Rate):").grid(row=1, column=0, sticky="nw")
    items_text = tk.Text(form_frame, height=6, width=50)
    items_text.grid(row=1, column=1, padx=5)

    ttk.Label(form_frame, text="Items (format: Description,Qty,Rate):").grid(row=1, column=0, sticky="nw")
    items_text = tk.Text(form_frame, height=6, width=50)
    items_text.grid(row=1, column=1, padx=5)

    ttk.Label(form_frame, text="Due Date:").grid(row=2, column=0, sticky="w")
    due_date = DateEntry(form_frame, date_pattern="yyyy-mm-dd")
    due_date.grid(row=2, column=1, padx=5, sticky="w")

    ttk.Label(form_frame, text="Template Style:").grid(row=3, column=0, sticky="w")
    style_var = tk.StringVar(value="Classic")
    style_dropdown = ttk.Combobox(form_frame, textvariable=style_var, values=["Classic", "Modern Bordered"], width=30)
    style_dropdown.grid(row=3, column=1, sticky="w", padx=5)

    def draw_logo(c):
        if os.path.exists(LOGO_PATH):
            try:
                from reportlab.lib.utils import ImageReader
                c.drawImage(ImageReader(LOGO_PATH), 400, 750, width=150, height=50, preserveAspectRatio=True)
            except Exception as e:
                print("Logo error:", e)

    def auto_suggest_items():
        client = client_entry.get().strip()
        if not client:
            messagebox.showwarning("Required", "Enter client name to auto-fill.")
            return
        if not os.path.exists(INVOICE_FILE):
            messagebox.showinfo("No Data", "No invoices found to suggest from.")
            return
        book = load_workbook(INVOICE_FILE)
        sheet = book.active
        past_items = [
            f"{row[1]},{row[2]},{row[3]}"
            for row in sheet.iter_rows(min_row=2, values_only=True)
            if row[0] == client
        ]
        if past_items:
            items_text.delete("1.0", tk.END)
            items_text.insert("1.0", "\n".join(past_items))
        else:
            messagebox.showinfo("No Match", f"No past items found for '{client}'")

    def save_invoice():
        client = client_entry.get().strip()
        items_lines = items_text.get("1.0", tk.END).strip().splitlines()
        due = due_date.get()
        style = style_var.get()
        if not client or not items_lines:
            messagebox.showerror("Missing", "Client and items are required.")
            return
        items = []
        total = 0
        for line in items_lines:
            try:
                desc, qty, rate = line.split(",")
                qty = int(qty)
                rate = float(rate)
                items.append((desc.strip(), qty, rate))
                total += qty * rate
            except:
                messagebox.showerror("Error", f"Invalid item line: {line}")
                return
        if not os.path.exists(INVOICE_FILE):
            wb = Workbook()
            ws = wb.active
            ws.append(["Client", "Description", "Qty", "Rate", "Total", "Due Date", "Status"])
        else:
            wb = load_workbook(INVOICE_FILE)
            ws = wb.active
        for desc, qty, rate in items:
            ws.append([client, desc, qty, rate, qty * rate, due, "Unpaid"])
        wb.save(INVOICE_FILE)

        # PDF
        c = canvas.Canvas(INVOICE_PDF, pagesize=A4)
        draw_logo(c)
        c.setFont("Helvetica-Bold", 14)
        c.drawString(30, 800, f"Invoice - {client}")
        c.setFont("Helvetica", 10)
        y = 770
        c.drawString(30, y, f"Due Date: {due}")
        y -= 30
        c.drawString(30, y, "Description")
        c.drawString(200, y, "Qty")
        c.drawString(260, y, "Rate")
        c.drawString(320, y, "Total")
        y -= 20
        for desc, qty, rate in items:
            c.drawString(30, y, desc)
            c.drawString(200, y, str(qty))
            c.drawString(260, y, f"{rate:.2f}")
            c.drawString(320, y, f"{qty * rate:.2f}")
            if style == "Modern Bordered":
                c.line(30, y - 2, 400, y - 2)
            y -= 20
        c.setFont("Helvetica-Bold", 12)
        c.drawString(30, y - 20, f"Grand Total: {total:.2f}")
        c.save()
        os.startfile(INVOICE_PDF)

        # HTML
        html = f"""
        <html><head><title>Invoice</title></head>
        <body style='font-family:sans-serif;'>
        <h2>Invoice - {client}</h2>
        <p><strong>Due Date:</strong> {due}</p>
        <table border='1' cellpadding='8' cellspacing='0'>
        <tr><th>Description</th><th>Qty</th><th>Rate</th><th>Total</th></tr>
        {''.join(f'<tr><td>{d}</td><td>{q}</td><td>{r}</td><td>{q*r}</td></tr>' for d, q, r in items)}
        <tr><td colspan='3'><strong>Grand Total</strong></td><td>{total}</td></tr>
        </table></body></html>"""
        with open(INVOICE_HTML, "w", encoding="utf-8") as f:
            f.write(html)

        messagebox.showinfo("Saved", "Invoice saved, PDF opened, and HTML created.")

    def view_invoices():
        win = tk.Toplevel()
        win.title("Invoice History")
        tree = ttk.Treeview(win, columns=("Client", "Description", "Qty", "Rate", "Total", "Due", "Status"), show="headings")
        for col in tree["columns"]:
            tree.heading(col, text=col)
            tree.column(col, width=100)
        tree.pack(fill="both", expand=True)

        if os.path.exists(INVOICE_FILE):
            book = load_workbook(INVOICE_FILE)
            sheet = book.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                tree.insert("", "end", values=row)

        def mark_paid():
            selected = tree.selection()
            if not selected:
                return
            index = tree.index(selected[0]) + 2
            book = load_workbook(INVOICE_FILE)
            sheet = book.active
            sheet[f"G{index}"].value = "Paid"
            book.save(INVOICE_FILE)
            win.destroy()
            view_invoices()

        def export_pdf():
            selected = tree.selection()
            if not selected:
                return
            row_data = tree.item(selected[0])["values"]
            client, desc, qty, rate, total, due, _ = row_data
            c = canvas.Canvas(INVOICE_PDF, pagesize=A4)
            draw_logo(c)
            c.setFont("Helvetica-Bold", 14)
            c.drawString(30, 800, f"Invoice - {client}")
            c.setFont("Helvetica", 10)
            c.drawString(30, 770, f"Due Date: {due}")
            c.drawString(30, 740, "Description")
            c.drawString(200, 740, "Qty")
            c.drawString(260, 740, "Rate")
            c.drawString(320, 740, "Total")
            c.drawString(30, 720, str(desc))
            c.drawString(200, 720, str(qty))
            c.drawString(260, 720, str(rate))
            c.drawString(320, 720, str(total))
            c.setFont("Helvetica-Bold", 12)
            c.drawString(30, 680, f"Grand Total: {total}")
            c.save()
            os.startfile(INVOICE_PDF)

        btn_frame = ttk.Frame(win)
        btn_frame.pack(fill="x", pady=5)
        ttk.Button(btn_frame, text="Mark as Paid", command=mark_paid).pack(side="left", padx=5)
        ttk.Button(btn_frame, text="Export PDF", command=export_pdf).pack(side="left", padx=5)

    ttk.Button(form_frame, text="🧠 Suggest Items", command=auto_suggest_items).grid(row=4, column=1, sticky="w", pady=5)
    ttk.Button(form_frame, text="Generate Invoice", command=save_invoice).grid(row=5, column=1, sticky="e", pady=10)
    ttk.Button(form_frame, text="View Invoice History", command=view_invoices).grid(row=6, column=1, sticky="e", pady=5)

# ---------------------------- Task Calendar ----------------------------
def load_tasks():
    if os.path.exists(TASKS_FILE):
        with open(TASKS_FILE, "r") as f:
            return [line.strip().split("|", 2) for line in f if line.count("|") >= 1]
    return []

def save_all_tasks(tasks):
    with open(TASKS_FILE, "w") as f:
        for date, text, *rest in tasks:
            category = rest[0] if rest else "General"
            f.write(f"{date}|{text}|{category}\n")

def save_task(task_date, task_text, category="General"):
    with open(TASKS_FILE, "a") as f:
        f.write(f"{task_date}|{task_text}|{category}\n")

def show_calendar_tab(notebook):
    calendar_tab = ttk.Frame(notebook)
    notebook.add(calendar_tab, text="Task Calendar")

    cal_frame = tk.Frame(calendar_tab)
    cal_frame.pack(side="left", fill="both", expand=True, padx=10, pady=10)

    cal = Calendar(cal_frame, selectmode='day', date_pattern='yyyy-mm-dd')
    cal.pack(fill="both", expand=True)

    day_tasks_frame = tk.Frame(cal_frame)
    day_tasks_frame.pack(fill="x")

    task_listbox = tk.Listbox(day_tasks_frame, height=6)
    task_listbox.pack(fill="x", padx=5, pady=2)

    controls_frame = tk.Frame(calendar_tab)
    controls_frame.pack(side="right", fill="y", padx=10, pady=10)

    task_entry = ttk.Entry(controls_frame, width=40)
    task_entry.pack(pady=5)

    category_var = tk.StringVar()
    category_combo = ttk.Combobox(controls_frame, textvariable=category_var, values=["General", "Follow-up", "Meeting", "Call"], width=38)
    category_combo.pack(pady=5)
    category_combo.set("General")

    def refresh_tasks():
        task_listbox.delete(0, tk.END)
        selected_date = cal.get_date()
        tasks = load_tasks()
        for d, t, *cat in tasks:
            if d == selected_date:
                display = f"{t} [{cat[0] if cat else 'General'}]"
                task_listbox.insert(tk.END, display)
        highlight_calendar_dates(cal)

    def add_task():
        task_text = task_entry.get().strip()
        category = category_var.get()
        if task_text:
            save_task(cal.get_date(), task_text, category)
            task_entry.delete(0, tk.END)
            refresh_tasks()

    def delete_task():
        selected = task_listbox.curselection()
        if selected:
            index = selected[0]
            selected_date = cal.get_date()
            selected_task = task_listbox.get(index).split(" [")[0]
            tasks = load_tasks()
            filtered = [(d, t, c) for d, t, *c in tasks if not (d == selected_date and selected_task == t)]
            save_all_tasks(filtered)
            refresh_tasks()

    def edit_task():
        selected = task_listbox.curselection()
        if selected:
            index = selected[0]
            current_display = task_listbox.get(index)
            current_text = current_display.split(" [")[0]
            current_cat = current_display.split("[")[-1].rstrip("]")
            new_text = simpledialog.askstring("Edit Task", "Edit your task:", initialvalue=current_text)
            new_cat = simpledialog.askstring("Edit Category", "Edit category:", initialvalue=current_cat)
            if new_text:
                selected_date = cal.get_date()
                tasks = load_tasks()
                updated_tasks = [(d, new_text if d == selected_date and t == current_text else t, new_cat if d == selected_date and t == current_text else (c[0] if c else "General")) for d, t, *c in tasks]
                save_all_tasks(updated_tasks)
                refresh_tasks()

    def highlight_calendar_dates(calendar):
        tasks = load_tasks()
        calendar.calevent_remove("all")
        date_text_map = {}
        for d, t, *c in tasks:
            if d not in date_text_map:
                date_text_map[d] = []
            date_text_map[d].append(f"- {t}")

        for date_str, task_lines in date_text_map.items():
            try:
                calendar.calevent_create(datetime.strptime(date_str, "%Y-%m-%d"), "\n".join(task_lines), 'task')
            except:
                pass

        calendar.tag_config('task', background='lightblue', foreground='black')

    ttk.Button(controls_frame, text="Add Task", command=add_task).pack(pady=2)
    ttk.Button(controls_frame, text="Delete Task", command=delete_task).pack(pady=2)
    ttk.Button(controls_frame, text="Edit Task", command=edit_task).pack(pady=2)
    cal.bind("<<CalendarSelected>>", lambda e: refresh_tasks())
    refresh_tasks()

# ---------------------------- Notifications ----------------------------
def show_task_reminders():
    today = date.today()
    upcoming = today + timedelta(days=1)
    tasks = load_tasks()
    reminder_tasks = [f"- {t} [{c[0] if c else 'General'}]" for d, t, *c in tasks if d == today.strftime("%Y-%m-%d") or d == upcoming.strftime("%Y-%m-%d")]
    if reminder_tasks:
        messagebox.showinfo("Upcoming Tasks", "\n".join(reminder_tasks))

def hourly_task_checker():
    while True:
        now = datetime.now()
        current_time = now.strftime("%H:%M")
        current_date = now.strftime("%Y-%m-%d")
        tasks = load_tasks()
        for d, t, *c in tasks:
            if d == current_date and current_time in t:
                messagebox.showinfo("Task Reminder", f"It's time for: {t}")
        time.sleep(60)

# ---------------------------- GUI Launcher ----------------------------

def show_task_reminders():
    today_str = date.today().strftime("%Y-%m-%d")
    tasks = load_tasks()
    today_tasks = [f"- {t} [{c[0] if c else 'General'}]" for d, t, *c in tasks if d == today_str]
    if today_tasks:
        messagebox.showinfo("Today's Tasks", "\n".join(today_tasks))

def start_crm():
    global result_tree, search_name_var, search_status_var, fields, notes_text, followup_date, selected_lead_index
    selected_lead_index = None

    root = tk.Tk()
    root.title("CRM - Look-Out Real Property")
    root.geometry("1200x700")
    root.configure(bg="#F9F9F9")
    notebook = ttk.Notebook(root)
    notebook.pack(fill="both", expand=True)

    style = ttk.Style()
    style.theme_use("clam")
    style.configure("TFrame", background="#F9F9F9")
    style.configure("TButton", background="#0078D7", foreground="white", font=("Segoe UI", 10), padding=6)
    style.configure("TLabel", background="#F9F9F9", font=("Segoe UI", 10))
    style.configure("TEntry", relief="flat")
    style.configure("Treeview", background="white", fieldbackground="white", rowheight=25, font=("Segoe UI", 9))
    style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"), background="#0078D7", foreground="white")

    notebook = ttk.Notebook(root)
    notebook.pack(fill="both", expand=True, padx=10, pady=10)

    tab1 = ttk.Frame(notebook)
    tab2 = ttk.Frame(notebook)
    notebook.add(tab1, text="Lead Manager")
    notebook.add(tab2, text="Reports")

    # Lead Manager Tab
    form_frame = ttk.Frame(tab1)
    form_frame.pack(side="top", fill="x", padx=10, pady=5)

    labels = ["Name", "Phone", "Email", "Source", "Property", "Status"]
    fields = {}
    for i, label in enumerate(labels):
        ttk.Label(form_frame, text=label).grid(row=i, column=0, sticky="w")
        entry = ttk.Entry(form_frame, width=30)
        entry.grid(row=i, column=1, padx=5, pady=3, sticky="w")
        fields[label] = entry


    ttk.Label(form_frame, text="Follow-up").grid(row=6, column=0, sticky="w")
    followup_date = DateEntry(form_frame, width=28, date_pattern="yyyy-mm-dd")
    followup_date.grid(row=6, column=1, padx=5, pady=3, sticky="w")

    ttk.Label(form_frame, text="Notes").grid(row=7, column=0, sticky="nw")
    notes_text = tk.Text(form_frame, height=4, width=40)
    notes_text.grid(row=7, column=1, padx=5, pady=3, sticky="w")

    btn_frame = ttk.Frame(tab1)
    btn_frame.pack(fill="x", padx=10, pady=5)
    actions = [
        ("Add Lead", add_lead), ("Load", load_selected_lead),
        ("Update", update_lead), ("Delete", delete_lead),
        ("Clear", clear_form), ("WhatsApp", send_whatsapp),
        ("Email", send_email)
    ]
    for i, (label, cmd) in enumerate(actions):
        ttk.Button(btn_frame, text=label, command=cmd).pack(side="left", padx=5)

    search_frame = ttk.Frame(tab1)
    search_frame.pack(fill="x", padx=10)
    ttk.Label(search_frame, text="Search Name:").pack(side="left")
    search_name_var = tk.StringVar()
    ttk.Entry(search_frame, textvariable=search_name_var, width=30).pack(side="left", padx=5)
    ttk.Label(search_frame, text="Status:").pack(side="left")
    search_status_var = tk.StringVar()
    status_filter = ttk.Combobox(search_frame, textvariable=search_status_var, values=["All", "Hot", "Warm", "Cold"], width=10)
    status_filter.pack(side="left", padx=5)
    status_filter.current(0)
    ttk.Button(search_frame, text="🔍 Search", command=search_leads).pack(side="left", padx=10)

    result_tree = ttk.Treeview(tab1, columns=("Name", "Phone", "Email", "Source", "Property", "Status", "Follow-up", "Notes"), show="headings")
    for col in result_tree["columns"]:
        result_tree.heading(col, text=col)
        result_tree.column(col, anchor="center")
    result_tree.pack(fill="both", expand=True, padx=10, pady=10)

    ttk.Button(btn_frame, text="Inactive Leads", command=lambda: show_inactive_leads(30)).pack(side="left", padx=5)

    # Add extra features: Follow-ups, Print Detail, Tagging
    ttk.Button(btn_frame, text="Today Follow-ups", command=show_followups_today).pack(side="left", padx=5)
    ttk.Button(btn_frame, text="Print Lead Detail", command=print_selected_lead_detail).pack(side="left", padx=5)
    create_tag_menu(btn_frame)



    # Reports Tab
    def generate_charts():
        for widget in tab2.winfo_children():
            widget.destroy()
        leads = get_leads()
        if not leads:
            ttk.Label(tab2, text="No leads to analyze.", font=("Segoe UI", 12)).pack(pady=20)
            return
        status_counts = Counter([l['Status'] for l in leads])
        fig, ax = plt.subplots(figsize=(5, 4))
        ax.bar(status_counts.keys(), status_counts.values(), color=['#E74C3C', '#F1C40F', '#3498DB'])
        ax.set_title("Lead Status Summary")
        ax.set_ylabel("Number of Leads")
        ax.set_xlabel("Status")
        chart = FigureCanvasTkAgg(fig, master=tab2)
        chart.draw()
        chart.get_tk_widget().pack(pady=30)
        ttk.Button(tab2, text="🔄 Refresh Chart", command=generate_charts).pack()

    generate_charts()
    search_leads()
    show_calendar_tab(notebook)
    show_invoice_tab(notebook)
    threading.Thread(target=hourly_task_checker, daemon=True).start()
    show_task_reminders()
    show_followups_today()
    root.mainloop()

start_crm()
